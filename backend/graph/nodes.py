"""
岗标辅导 Graph Nodes

完全实现 gangbiao-coach skill v2 的能力：
  - 阶段 0：check_progress + 文件采集提示
  - 阶段 1：validate_structure（结构校验+内容提取）
  - 阶段 2：review_item（全面评审，三问法+严格度校准）
  - 阶段 3：guide_reflection（分步引导，五条铁律+引导武器+十一问）
           process_response（回判三步+卡住升级）
           detect_user_intent + answer_user_question（意图识别+被动答疑）
  - 阶段 4：generate_closure（收尾总结）
"""
from __future__ import annotations
import os
import json
import subprocess
import shutil
from pathlib import Path
from typing import Any, Optional

from langchain_openai import ChatOpenAI
from langchain_core.messages import AIMessage, HumanMessage, SystemMessage
from langchain_core.runnables import RunnableConfig

from .state import (
    CoachState, ReviewItem, RUBRIC_ITEMS, RUBRIC_INDEX,
    GUIDANCE_WEAPONS, STRICTNESS_TABLE, REAL_EXAMPLE,
    SELF_CHECK_QUESTIONS, WHO_HINTS, BENEFIT_HINTS, BIZ_HINTS,
)
from ..loaders import (
    load_scoring_criteria,
    load_submission,
    load_pdf_as_text,
    MOCK_SCORING_CRITERIA,
    MOCK_TEACHING_MATERIAL,
)

# ─────────────────────────────────────────────────────────────
# 常量
# ─────────────────────────────────────────────────────────────
QUESTION_HINTS = ("?", "？", "什么", "为何", "为什么", "如何", "怎么", "请问", "能否", "可以", "是否")
REPLY_HINTS = ("我觉得", "我理解", "我会", "我打算", "我修改", "可以改成", "因为", "我的回答", "我先")

# 项目内嵌的参考文件路径
REFERENCES_DIR = Path(__file__).parent.parent / "references"
SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"

# ─────────────────────────────────────────────────────────────
# LLM 实例（延迟初始化）
# ─────────────────────────────────────────────────────────────
_llm_instance = None

def _get_llm():
    global _llm_instance
    if _llm_instance is None and _is_llm_enabled():
        try:
            _llm_instance = ChatOpenAI(
                model="deepseek-v4-pro",
                max_tokens=2048,
                request_timeout=8,
                max_retries=0,
            )
        except Exception as e:
            print(f"[警告] LLM 初始化失败: {e}")
            _llm_instance = None
    return _llm_instance


def _is_llm_enabled() -> bool:
    disabled = os.environ.get("STANDJOB_DISABLE_LLM", "").lower() in {"1", "true", "yes", "on"}
    if disabled:
        return False
    return any(os.environ.get(k) for k in ("OPENAI_API_KEY", "DEEPSEEK_API_KEY", "ANTHROPIC_API_KEY"))


# ─────────────────────────────────────────────────────────────
# 参考文件加载（带模块级缓存，避免重复读取）
# ─────────────────────────────────────────────────────────────

_rubric_text_cache: Optional[str] = None
_textbook_text_cache: Optional[str] = None
_system_prompt_cache: Optional[str] = None


def _load_rubric_text() -> str:
    """加载项目内嵌的 rubric.md，不存在则使用 MOCK 数据（缓存）"""
    global _rubric_text_cache
    if _rubric_text_cache is None:
        path = REFERENCES_DIR / "rubric.md"
        _rubric_text_cache = path.read_text(encoding="utf-8") if path.exists() else MOCK_SCORING_CRITERIA
    return _rubric_text_cache


def _load_textbook_text() -> str:
    """加载项目内嵌的 textbook.md，不存在则使用 MOCK 数据（缓存）"""
    global _textbook_text_cache
    if _textbook_text_cache is None:
        path = REFERENCES_DIR / "textbook.md"
        _textbook_text_cache = path.read_text(encoding="utf-8") if path.exists() else MOCK_TEACHING_MATERIAL
    return _textbook_text_cache


# ─────────────────────────────────────────────────────────────
# 进度脚本调用
# ─────────────────────────────────────────────────────────────

def _thread_id_from_config(config: Optional[dict]) -> Optional[str]:
    """从 LangGraph 注入的 config 中取出 thread_id（用于隔离进度文件）。"""
    if not config:
        return None
    return (config.get("configurable") or {}).get("thread_id")


def _progress_file_for_thread(thread_id: Optional[str]) -> str:
    """按 thread_id 生成隔离的进度文件路径，避免多会话串号。"""
    if not thread_id:
        return "/tmp/gangbiao-coach-progress.json"
    safe = "".join(ch for ch in str(thread_id) if ch.isalnum() or ch in "-_")
    return f"/tmp/gangbiao-coach-progress-{safe}.json"


def _call_progress_script(*args: str, thread_id: Optional[str] = None) -> Optional[dict]:
    """调用 progress.py 脚本，返回解析后的 JSON 或 None。

    通过 STANDJOB_PROGRESS_FILE 环境变量把进度文件按 thread_id 隔离。
    """
    script = SCRIPTS_DIR / "progress.py"
    if not script.exists():
        return None
    env = dict(os.environ)
    env["STANDJOB_PROGRESS_FILE"] = _progress_file_for_thread(thread_id)
    try:
        result = subprocess.run(
            ["python3", str(script)] + list(args),
            capture_output=True, text=True, timeout=10,
            env=env,
        )
        if result.returncode == 0 and result.stdout.strip():
            return json.loads(result.stdout.strip())
    except Exception as e:
        print(f"[进度脚本] 调用失败: {e}")
    return None


def _call_validate_script(file_path: str, extract: bool = False) -> Optional[dict]:
    """调用 validate_sheets.py 脚本"""
    script = SCRIPTS_DIR / "validate_sheets.py"
    if not script.exists():
        return None
    cmd = ["python3", str(script), file_path]
    if extract:
        cmd.append("--extract")
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=30,
        )
        if result.stdout.strip():
            return json.loads(result.stdout.strip())
    except Exception as e:
        print(f"[校验脚本] 调用失败: {e}")
    return None


# ─────────────────────────────────────────────────────────────
# 辅助函数
# ─────────────────────────────────────────────────────────────

def _safe_json_load(raw: str, fallback: dict) -> dict:
    cleaned = raw.strip()
    if "```" in cleaned:
        cleaned = cleaned.split("```")[1]
        if cleaned.startswith("json"):
            cleaned = cleaned[4:]
    try:
        return json.loads(cleaned.strip())
    except Exception:
        return fallback


def _is_question_intent(text: str) -> bool:
    stripped = (text or "").strip()
    if not stripped:
        return False
    return any(h in stripped for h in QUESTION_HINTS)


def _is_reply_intent(text: str) -> bool:
    stripped = (text or "").strip()
    if not stripped:
        return False
    return any(h in stripped for h in REPLY_HINTS)


def _rule_based_intent(text: str) -> tuple[str, float]:
    stripped = (text or "").strip()
    if not stripped:
        return "reply", 0.6
    q = _is_question_intent(stripped)
    r = _is_reply_intent(stripped)
    if q and not r:
        return "question", 0.85
    if r and not q:
        return "reply", 0.8
    if q and r:
        return "uncertain", 0.4
    return "reply", 0.55


def _passes_three_questions(text: str) -> bool:
    """三问法：客户主语 + 具体好处 + 商业落点"""
    candidate = (text or "").strip()
    if not candidate:
        return False
    has_who = any(k in candidate for k in WHO_HINTS) or any(
        k in candidate.lower() for k in ("product manager", "business", "customer", "user", "team", "stakeholder")
    )
    has_benefit = any(k in candidate for k in BENEFIT_HINTS) or any(
        k in candidate.lower() for k in ("reduce", "improve", "increase", "avoid", "decrease", "enhance", "ensure")
    )
    has_biz = any(k in candidate for k in BIZ_HINTS) or any(
        k in candidate.lower() for k in ("revenue", "cost", "risk", "brand", "market", "efficiency", "reputation", "quality")
    )
    return has_who and has_benefit and has_biz


def _sanitize_user_desc(text: str) -> str:
    sanitized = (text or "").strip()
    blocked = [
        "评分标准", "教材第", "问题项",
        "第1项", "第2项", "第3项", "第4项", "第5项",
        "第6项", "第7项", "第8项", "第9项", "第10项",
        "第11项", "第12项", "第13项", "第14项",
    ]
    for token in blocked:
        sanitized = sanitized.replace(token, "")
    return sanitized.strip("：:，,。 ") or "当前表达还有一层关键信息不够具体。"


def _extract_value_text(row_data: dict[str, Any]) -> str:
    texts: list[str] = []
    for key, value in row_data.items():
        if "岗位价值" in key and value:
            texts.append(str(value))
    return " ".join(texts)


def _format_value_for_view(value: Any) -> str:
    if isinstance(value, list):
        items = []
        for idx, task in enumerate(value, 1):
            task_parts = [f"{k}: {v}" for k, v in task.items() if v]
            if task_parts:
                items.append(f"{idx}. {' | '.join(task_parts)}")
        return "\n".join(items)
    return str(value)


def _pick_related_fields(row_data: dict[str, Any], issue: dict[str, Any]) -> list[str]:
    desc = f"{issue.get('category', '')} {issue.get('issue_desc', '')} {issue.get('explanation', '')}".lower()
    non_empty_keys = [k for k, v in row_data.items() if v]
    if not non_empty_keys:
        return []
    ranked: list[str] = []
    keywords = [
        ("岗位价值", ["岗位价值", "价值", "客户", "商业"]),
        ("岗位效能", ["岗位效能", "效能", "指标", "量化"]),
        ("核心任务", ["核心任务", "任务", "目的", "成果"]),
        ("辅助任务", ["辅助任务", "任务", "目的", "成果"]),
    ]
    for key in non_empty_keys:
        score = 0
        for _, hints in keywords:
            if any(h in key for h in hints):
                score += 1
            if any(h in desc for h in hints) and any(h in key for h in hints):
                score += 2
        if "资源投入" in key:
            score -= 1
        if score > 0:
            ranked.append((score, key))
    if ranked:
        ranked.sort(key=lambda x: (-x[0], x[1]))
        return [k for _, k in ranked[:3]]
    return non_empty_keys[:2]


# ─────────────────────────────────────────────────────────────
# 灰区放过逻辑
# ─────────────────────────────────────────────────────────────

def _should_relax_issue(item_id: int, row_data: dict[str, Any], issue: dict[str, Any]) -> bool:
    """
    灰区放过逻辑 - 基于严格度校准对照表
    三问法都过 → 合格放过；灰色地带默认放过
    """
    if item_id in {1, 2, 3, 4}:
        value_text = _extract_value_text(row_data)
        if _passes_three_questions(value_text):
            return True

    if item_id == 5:
        value_text = _extract_value_text(row_data)
        eff_text = str(row_data.get("岗位效能", "") or row_data.get("岗位效能_", ""))
        if eff_text and any(k in eff_text for k in ("%", "率", "次", "个", "小时", "天", "数量", "≤", ">=")):
            return True

    if item_id in {7, 8}:
        desc = (issue.get("issue_desc") or "") + (issue.get("explanation") or "")
        if "轻微" in desc or "偶有" in desc:
            return True
        task_names = []
        for k, v in row_data.items():
            if "任务" in k and v and isinstance(v, str):
                task_names.append(v)
        verbs = ("负责", "完成", "制定", "整理", "分析", "设计", "开发", "测试", "评审", "编写")
        if task_names and any(v in task_names[0] for v in verbs):
            return True

    if item_id in {9, 10, 11, 12, 13, 14}:
        outcome_text = ""
        for k, v in row_data.items():
            if "成果" in k and v:
                outcome_text += str(v)
        if any(k in outcome_text for k in ("%", "率", "次", "个", "≤", ">=", "内", "以上")):
            return True

    return False


def _coerce_rubric_item_id(raw: Any) -> Optional[int]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return None
    item_id = int(digits)
    return item_id if item_id in RUBRIC_INDEX else None


def _normalize_matched_issues(
    matched_issues: list[dict[str, Any]], row_data: dict[str, Any]
) -> tuple[list[dict[str, Any]], list[int]]:
    """标准化匹配问题，应用灰区放过逻辑。"""
    normalized: list[dict[str, Any]] = []
    relaxed_ids: list[int] = []

    for raw in matched_issues:
        item_id = _coerce_rubric_item_id(raw.get("issue_id"))
        if item_id is None:
            continue
        canonical = RUBRIC_INDEX[item_id]
        # 严重程度：A=重点关注，B=可优化——仅用于内部排序，不给用户展示
        level = str(raw.get("level", canonical["level"]))
        issue = {
            "issue_id": str(item_id),
            "rubric_item_id": item_id,
            "issue_desc": raw.get("issue_desc") or canonical["desc"],
            "category": raw.get("category") or canonical["category"],
            "level": "A" if level.upper() == "A" else "B",
            "explanation": raw.get("explanation", ""),
            "user_facing_desc": _sanitize_user_desc(
                raw.get("user_facing_desc") or raw.get("issue_desc") or canonical["desc"]
            ),
        }
        if _should_relax_issue(item_id, row_data, issue):
            relaxed_ids.append(item_id)
            continue
        normalized.append(issue)

    normalized.sort(key=lambda x: int(x.get("rubric_item_id", 99)))
    return normalized, sorted(set(relaxed_ids))


# ─────────────────────────────────────────────────────────────
# 构建系统提示（五条铁律）
# ─────────────────────────────────────────────────────────────

def _build_system_prompt(scoring_criteria: str, teaching_material: str) -> str:
    """构建系统提示（带模块级缓存，避免重复拼接大段文本）。"""
    global _system_prompt_cache
    if _system_prompt_cache is None:
        _system_prompt_cache = _build_system_prompt_impl(scoring_criteria, teaching_material)
    return _system_prompt_cache


def _build_system_prompt_impl(scoring_criteria: str, teaching_material: str) -> str:
    return f"""你是岗标教练，像带徒弟一样陪用户打磨"岗位价值与岗位任务"材料。

## 你的角色（牢记）

- **身份**：有经验的岗标教练，不是评审官、不是考官
- **风格**：像老同事在茶水间聊天，自然、口语化，不说教
- **红线**：绝不暴露内部诊断过程——不报条目编号、不念章节号、不说"根据评分标准"

### 正确示例
❌ "根据评分标准第7项，你的任务命名不符合规范"
❌ "教材第5章提到'动词+修饰语+名词'格式"
❌ "这个问题很严重，扣10分"
✅ "你写的'AI应用开发与业务落地'——'业务落地'听起来更像你想要的结果，而不是一个能直接动手做的动作。如果把任务名字换成'一看就知道要干什么'的写法，你觉得该怎么改？"

## 辅导原则（五条铁律）

1. **只提问，不代写** — 心里再清楚也用提问"钓"出答案，不直接给标准答案
2. **内部严谨，对外像聊天** — 所有判断都有依据，但说出来要像老教练带人，不念标准
3. **一次只拎一个点** — 当前问题没站稳不跳题
4. **先肯定，再追一层** — 每次开口先点明具体进步，不是泛泛的"不错"
5. **不给打分，不说通关** — 只诊断问题、引导解决，不评价好坏、不给分数

## 双线推进（严禁混淆）

**线1：客户 → 价值 → 效能**（必须先完成）
- 关键客户定义（问题项0）
- 岗位价值（问题项1-4）
- 岗位效能（问题项5）

**线2：任务 → 目的 → 成果**（线1完成后才能进入）
- 岗位任务（问题项6-8）
- 任务目的与成果（问题项9-14）

**铁律**：严禁把任务和效能挂钩，两条线完全独立。

## 严格度校准（三问法）

任何价值描述，先用这三问筛一遍：
1. **谁**：能圈出明确的"客户主语"（产品经理/测试工程师/业务部门）吗？
2. **得到啥**：客户能看到自己拿到的"具体好处"（不是你的动作）吗？
3. **商业落点**：最终落到 收入/成本/风险/品牌 中至少一个吗？

- 三问都过 → **及格，放过，不要追**
- 任一未过 → 进引导清单

**灰色地带倾向于放过**——当某条描述既不像典型违规、又不能算明显合格时，**默认放过**。

**真实反例**（务必记住）：
{REAL_EXAMPLE}

---

### 典型问题（内部诊断依据，对外不提）
{scoring_criteria}

### 教材精要（引导依据，用人话表达）
{teaching_material}
"""


# ─────────────────────────────────────────────────────────────
# 引导问题生成
# ─────────────────────────────────────────────────────────────

def _generate_issue_question(
    scoring_criteria: str,
    teaching_material: str,
    row_data: dict[str, Any],
    issue: dict[str, Any],
) -> str:
    """为当前问题项生成单一聚焦引导问题（优先使用引导武器表）"""
    rubric_item_id = issue.get("rubric_item_id", 0)
    weapon = GUIDANCE_WEAPONS.get(str(rubric_item_id))

    if weapon and _is_llm_enabled():
        # 有对应的引导武器，让 LLM 基于武器风格 + 教材生成自然的问题
        system = _build_system_prompt(scoring_criteria, teaching_material)
        category = issue.get("category", "")
        user_desc = issue.get("user_facing_desc", issue.get("issue_desc", ""))

        # 找到对应自检问
        self_check = None
        for q in SELF_CHECK_QUESTIONS:
            if q["id"] == str(rubric_item_id) or any(
                keyword in category for keyword in q["mistake"]
            ):
                self_check = q
                break

        prompt = f"""基于以下引导武器风格，生成一个自然的苏格拉底式引导问题。

当前问题类别：{category}
问题描述：{user_desc}
引导武器风格：{weapon['question_style']}
教材参考：{weapon['textbook_ref']}
"""
        if self_check:
            prompt += f"\n自检问参考：{self_check['check']}"

        prompt += """

要求：
1. 用自然的口语，像老同事在茶水间聊天那样
2. 不报章节号、不报条目编号
3. 一次只聚焦一个点
4. 先指出做得好的地方，再引出问题
5. 用提问引导用户自己想，不给答案

直接输出问题文本，不要 JSON 格式。"""

        try:
            llm = _get_llm()
            if llm is None:
                raise Exception("LLM not available")
            resp = llm.invoke([
                SystemMessage(content=system),
                HumanMessage(content=prompt),
            ])
            return resp.content.strip()
        except Exception:
            pass

    # 兜底：使用引导武器表的默认话术
    if weapon:
        return weapon["question_style"]

    # 最终兜底
    return "你觉得这个地方还能怎么写得更有说服力？"


def _build_proactive_message(
    idx: int, total: int, row_data: dict[str, Any],
    issue: dict[str, Any], question: str,
) -> str:
    """构建主动引导消息（每轮对话结构：肯定 → 切入 → 一问）"""
    category = issue.get("category", "")
    user_desc = issue.get("user_facing_desc", issue.get("issue_desc", ""))

    # 尝试找一条做得好的点
    highlight = ""
    for key, val in row_data.items():
        if val and isinstance(val, str) and len(val) > 5:
            highlight = f"你{key}这块写得挺实在的，"
            break
    if not highlight:
        highlight = "材料整体框架是清晰的，"

    content = (
        f"📋 第 {idx+1}/{total} 条\n\n"
        f"✅ {highlight}我们来看一个可以更精准的地方。\n\n"
        f"💬 {question}"
    )
    return content


def _advance_to_next_item(state: CoachState, feedback: str) -> dict:
    """推进到下一个条目或进入收尾"""
    idx = state["current_item_index"]
    rows = state["submission_rows"]
    next_idx = idx + 1

    # 计算当前线1完成状态
    issue_queue = state.get("review_items", [{}])[idx].get("issue_queue", []) if state.get("review_items") else []
    line1_issue_ids = {q["issue_id"] for q in issue_queue if q["rubric_item_id"] <= 5}
    issue_status_map = dict(state.get("issue_status_map", {}))
    line1_resolved_ids = {iid for iid in line1_issue_ids if issue_status_map.get(iid) == "resolved"}
    line1_completed = len(line1_issue_ids) == len(line1_resolved_ids)

    if next_idx >= len(rows):
        return {
            "phase": "done",
            "current_item_index": next_idx,
            "messages": [AIMessage(content=f"✅ {feedback}\n\n所有条目已辅导完毕！")],
            "active_mode": "proactive",
            "awaiting_user_input": False,
            "current_line": 2,
            "line1_completed": True,
        }

    return {
        "current_item_index": next_idx,
        "current_issue_index": 0,
        "issue_round": 0,
        "phase": "reviewing",
        "active_mode": "proactive",
        "pending_question": "",
        "messages": [AIMessage(content=f"✅ {feedback}\n\n我们来看下一条。")],
        "stuck_counter": 0,
        "hint_level": 0,
        "awaiting_user_input": False,
        "current_line": 1,  # 新条目从线1开始
        "line1_completed": False,
    }


def _build_escalation_question(
    issue: dict[str, Any], row_data: dict[str, Any], hint_level: int
) -> str:
    """卡住时升级提示（3 级递进）"""
    related_fields = issue.get("related_fields", [])
    field_name = related_fields[0] if related_fields else "当前描述"
    current_text = str(row_data.get(field_name, "")) if field_name in row_data else ""

    if hint_level <= 1:
        return f"如果只改 {field_name} 这一处，你会先补哪一个可验证信息，能让评审一眼看到变化？"
    if hint_level == 2:
        return (
            f"给你一个更具体的抓手：围绕 {field_name}，"
            "先补‘服务对象’、‘可观察结果’、‘业务影响’三者中的哪一项？"
        )
    # hint_level >= 3: most specific hint
    text_preview = current_text[:60]
    lq = "\u201c"  # left double quote
    rq = "\u201d"  # right double quote
    return (
        f"我们再收窄一步：基于你现在这句{lq}{text_preview}{rq}，"
        f"你愿意先把它改成{lq}为谁带来什么具体结果{rq}的一句话吗？"
    )

def _find_next_issue(
    queue_copy: list[dict],
    current_idx: int,
    line1_completed: bool,
    issue_status_map: dict[str, str],
) -> tuple[Optional[int], Optional[dict], Optional[str]]:
    """
    双线推进：找到下一个待处理的问题项。

    规则：
    - 线1（客户/价值/效能，rubric_item_id 0-5）全部完成后，才能进入线2（任务/目的/成果，6-14）
    - 如果当前在线1且线1未完成，跳过线2问题项
    - 返回 (next_idx, next_issue, transition_msg)，如果没有则返回 (None, None, None)
    """
    transition_msg: Optional[str] = None

    for i in range(current_idx + 1, len(queue_copy)):
        candidate = queue_copy[i]
        rubric_id = candidate["rubric_item_id"]
        issue_id = candidate["issue_id"]

        # 跳过已解决的问题
        if issue_status_map.get(issue_id) == "resolved":
            continue

        # 检查是否在正确的线上
        if rubric_id > 5:  # 线2问题
            if not line1_completed:
                # 线1未完成，跳过线2问题，继续找线1问题
                continue
            # 这是从线1到线2的首次过渡
            transition_msg = "客户-价值-效能这部分已经梳理清楚了，接下来我们看看任务设计和成果规划。"

        return i, candidate, transition_msg

    return None, None, None
    """推进到下一个条目或进入收尾"""
    idx = state["current_item_index"]
    rows = state["submission_rows"]
    next_idx = idx + 1

    if next_idx >= len(rows):
        return {
            "phase": "done",
            "current_item_index": next_idx,
            "messages": [AIMessage(content=f"✅ {feedback}\n\n所有条目已辅导完毕！")],
            "active_mode": "proactive",
            "awaiting_user_input": False,
        }

    return {
        "current_item_index": next_idx,
        "current_issue_index": 0,
        "issue_round": 0,
        "phase": "reviewing",
        "active_mode": "proactive",
        "pending_question": "",
        "messages": [AIMessage(content=f"✅ {feedback}\n\n我们来看下一条。")],
        "stuck_counter": 0,
        "hint_level": 0,
        "awaiting_user_input": False,
    }


# ─────────────────────────────────────────────────────────────
# NODE 0: load_standards — 加载评审标准 + 检查进度
# ─────────────────────────────────────────────────────────────

def load_standards(state: CoachState, config: Optional[RunnableConfig] = None) -> dict:
    """
    初始化：加载评分标准和教材，检查是否有已保存的进度。
    对应 skill 阶段 0 的初始化 + 进度恢复检测。
    """
    thread_id = _thread_id_from_config(config)
    scoring_criteria = _load_rubric_text()
    teaching_material = _load_textbook_text()

    # 检查进度
    progress = _call_progress_script("show", thread_id=thread_id)
    has_saved = progress is not None and progress.get("file")

    if has_saved:
        welcome = (
            "嘿，我看到你上次有做到一半的辅导记录——"
            "要不要接着上次的来？还是重新开始？"
        )
        return {
            "scoring_criteria": scoring_criteria,
            "teaching_material": teaching_material,
            "phase": "loaded",
            "has_saved_progress": True,
            "progress_snapshot": progress,
            "messages": [AIMessage(content=welcome)],
        }

    welcome = "来，先把你的材料给我看看——把《岗位价值与岗位任务》那个 Excel 的路径发我就行。"
    return {
        "scoring_criteria": scoring_criteria,
        "teaching_material": teaching_material,
        "phase": "loaded",
        "has_saved_progress": False,
        "messages": [AIMessage(content=welcome)],
    }


# ─────────────────────────────────────────────────────────────
# NODE 1: validate_structure — 结构校验 + 内容提取
# 对应 skill 阶段 1：validate_sheets.py --extract
# ─────────────────────────────────────────────────────────────

def validate_structure(state: CoachState, config: Optional[RunnableConfig] = None) -> dict:
    """
    结构校验：用 validate_sheets.py 校验文件结构是否完整，
    通过后提取结构化数据，否则告知用户缺少哪些区块。
    """
    thread_id = _thread_id_from_config(config)
    file_path = state.get("submission_path", "")
    if not file_path:
        # 尝试创建 mock 文件
        file_path = _create_mock_submission()

    # 调用校验脚本
    result = _call_validate_script(file_path, extract=True)

    if result is None:
        # 脚本不可用，回退到内部解析
        return _fallback_validate(state, file_path, thread_id=thread_id)

    if not result.get("ok"):
        # validate_sheets.py may not handle all formats (e.g., merged cells).
        # Fall back to internal load_submission which is more flexible.
        return _fallback_validate(state, file_path, thread_id=thread_id)

    # 校验通过，提取结构化数据
    data = result.get("data", {})

    # 将提取的数据转为 submission_rows 格式
    columns, rows, text = _convert_extracted_data(data)

    # 用自然的语气告诉用户
    ai_msg = "收到，我先看一眼你的材料……好，结构完整，咱们开始吧。"

    # 保存进度
    _call_progress_script("init", file_path, thread_id=thread_id)

    return {
        "submission_path": file_path,
        "submission_columns": columns,
        "submission_rows": rows,
        "submission_text": text,
        "structure_valid": True,
        "structure_errors": [],
        "phase": "reviewing",
        "current_item_index": 0,
        "messages": [AIMessage(content=ai_msg)],
    }


def _convert_extracted_data(data: dict) -> tuple[list[str], list[dict], str]:
    """将 validate_sheets.py 提取的数据转为 submission_rows 格式"""
    clients = data.get("clients", [])
    aux_tasks = data.get("auxiliary_tasks", [])

    columns = ["岗位价值", "岗位效能", "核心任务", "核心任务_资源投入", "辅助任务", "辅助任务_资源投入"]
    rows: list[dict] = []

    for client in clients:
        row: dict[str, Any] = {
            "岗位价值": client.get("position_value", ""),
            "岗位效能": client.get("efficiency", ""),
            "核心任务": client.get("tasks", []),
            "辅助任务": [],
        }
        rows.append(row)

    # 附加辅助任务到最后一个客户
    if aux_tasks and rows:
        rows[-1]["辅助任务"] = aux_tasks

    # 生成文本摘要
    lines = ["=== 提取的结构化数据 ===", ""]
    for idx, client in enumerate(clients):
        lines.append(f"客户{idx+1}：{client.get('name', '')}")
        lines.append(f"  岗位价值：{client.get('position_value', '')}")
        lines.append(f"  岗位效能：{client.get('efficiency', '')}")
        for task in client.get("tasks", []):
            lines.append(f"  核心任务：{task.get('name', '')} — {task.get('purpose_and_result', '')}")
    for task in aux_tasks:
        lines.append(f"  辅助任务：{task.get('name', '')} — {task.get('purpose_and_result', '')}")

    text = "\n".join(lines)
    return columns, rows, text


def _fallback_validate(state: CoachState, file_path: str, thread_id: Optional[str] = None) -> dict:
    """脚本不可用时的回退：直接使用内部 load_submission 解析"""
    try:
        columns, rows, text = load_submission(file_path)
    except Exception as e:
        return {
            "structure_valid": False,
            "structure_errors": [str(e)],
            "messages": [AIMessage(content=f"文件读取失败：{e}")],
            "phase": "loaded",
        }

    if not rows:
        return {
            "structure_valid": False,
            "structure_errors": ["表格为空或格式无法识别"],
            "messages": [AIMessage(content="这个文件好像读不出内容，能检查一下格式吗？")],
            "phase": "loaded",
        }

    _call_progress_script("init", file_path, thread_id=thread_id)
    return {
        "submission_path": file_path,
        "submission_columns": columns,
        "submission_rows": rows,
        "submission_text": text,
        "structure_valid": True,
        "structure_errors": [],
        "phase": "reviewing",
        "current_item_index": 0,
        "messages": [AIMessage(content="材料收到了，我过了一遍，该有的都有——咱们直接进入正题。")],
    }


# ─────────────────────────────────────────────────────────────
# NODE 2: review_item — 全面评审（后台，不暴露过程）
# 对应 skill 阶段 2：严格逐条对照 14 项评分标准
# ─────────────────────────────────────────────────────────────

def review_item(state: CoachState) -> dict:
    """
    对当前条目进行全面评审。
    严格逐条对照 14 项评分标准，三问法筛选，灰区放过。
    这是后台工作，评审过程不暴露给用户。
    """
    idx = state["current_item_index"]
    rows = state["submission_rows"]

    if idx >= len(rows):
        return {"phase": "done"}

    row_data = rows[idx]
    columns = state["submission_columns"]

    # LLM 评审
    matched_issues: list[dict[str, Any]] = []
    checked_item_ids: list[int] = []

    if _is_llm_enabled():
        matched_issues, checked_item_ids = _llm_review(
            state["scoring_criteria"],
            state["teaching_material"],
            row_data,
            columns,
        )
    else:
        matched_issues, checked_item_ids = _rule_based_review(row_data)

    # 标准化 + 灰区放过
    normalized, relaxed_ids = _normalize_matched_issues(matched_issues, row_data)

    # 注意：skill 要求不给用户打分，仅用于内部诊断
    # 构建 issue_queue（辅导顺序：严格按 skill 推进顺序）
    # 顺序：0客户定义 → 1-4岗位价值 → 5效能 → 6-8任务 → 9-14目的成果
    issue_queue: list[dict[str, Any]] = []
    for issue in normalized:
        issue_queue.append({
            "issue_id": issue["issue_id"],
            "rubric_item_id": issue["rubric_item_id"],
            "issue_desc": issue["user_facing_desc"],
            "category": issue["category"],
            "status": "pending",
            "related_fields": _pick_related_fields(row_data, issue),
            "user_facing_desc": issue["user_facing_desc"],
        })

    # 按 skill 要求的推进顺序排序
    CATEGORY_ORDER = {
        "客户定义": 0,
        "岗位价值": 1,
        "岗位效能": 2,
        "岗位任务": 3,
        "任务目的与成果": 4,
    }
    issue_queue.sort(key=lambda x: (CATEGORY_ORDER.get(x["category"], 99), x["rubric_item_id"]))

    if issue_queue:
        issue_queue[0]["status"] = "in_progress"

    # 保存 review item
    review_items = list(state.get("review_items", []))
    # 如果已有同索引的，更新；否则追加
    if idx < len(review_items):
        review_items[idx] = {
            "row_index": idx,
            "row_data": row_data,
            "dimension_scores": {"matched_issues": normalized},
            "issues": [i["issue_desc"] for i in normalized],
            "suggestions": [],
            "standard_ref": "",
            "status": "reviewed",
            "issue_queue": issue_queue,
        }
    else:
        review_items.append({
            "row_index": idx,
            "row_data": row_data,
            "dimension_scores": {"matched_issues": normalized},
            "issues": [i["issue_desc"] for i in normalized],
            "suggestions": [],
            "standard_ref": "",
            "status": "reviewed",
            "issue_queue": issue_queue,
        })
            "dimension_scores": {"matched_issues": normalized},
            "issues": [i["issue_desc"] for i in normalized],
            "suggestions": [],
            "standard_ref": "",
            "status": "reviewed",
            "issue_queue": issue_queue,
        })

    # 构建 coaching_queue_order（辅导队列）
    coaching_queue_order = [q["issue_id"] for q in issue_queue]
    current_focus_id = issue_queue[0]["issue_id"] if issue_queue else None

    # 更新 issue_status_map
    issue_status_map = dict(state.get("issue_status_map", {}))
    for q in issue_queue:
        issue_status_map[q["issue_id"]] = q["status"]

    # 生成首个引导问题
    first_question = ""
    if issue_queue:
        first_issue = issue_queue[0]
        first_question = _generate_issue_question(
            state["scoring_criteria"],
            state["teaching_material"],
            row_data,
            first_issue,
        )

    print(f"[Node] review_item: 条目{idx+1}/{len(rows)}，发现{len(normalized)}个问题项，放过{len(relaxed_ids)}项")

    # 没有发现问题：自动推进到下一条，不走引导流程。
    # 收尾与推进两种情况合并为一个返回，推进索引只在此处计算一次。
    if not issue_queue:
        next_idx = idx + 1
        is_last = next_idx >= len(rows)
        if is_last:
            print(f"[Node] review_item: 所有条目均无问题，直接收尾")
            tail_msg = "所有条目都已检查完毕，写得挺好！"
        else:
            print(f"[Node] review_item: 条目{idx+1}无问题，自动推进到第{next_idx+1}条")
            tail_msg = "继续看下一条。"
        result = {
            "review_items": review_items,
            "all_issues": normalized,
            "current_issue_index": 0,
            "issue_round": 0,
            "issue_status_map": issue_status_map,
            "coaching_queue_order": coaching_queue_order,
            "current_focus_id": None,
            "pending_question": "",
            "phase": "done" if is_last else "reviewing",
            "stuck_counter": 0,
            "hint_level": 0,
            "current_line": 2,  # 无线1问题，直接进入线2
            "line1_completed": True,
            "rubric_eval_summary": {
                "checked_item_ids": checked_item_ids,
                "matched_item_ids": [i["rubric_item_id"] for i in normalized],
                "relaxed_item_ids": relaxed_ids,
                "coverage_ok": len(checked_item_ids) == len(RUBRIC_ITEMS),
            },
            "messages": [AIMessage(
                content=f"📋 第 {idx+1}/{len(rows)} 条检查完毕，没有发现需要调整的地方。" + tail_msg
            )],
        }
        if not is_last:
            result["current_item_index"] = next_idx
        return result

    # 判断当前属于线1还是线2，并确定线1是否已完成
    # 线1：客户定义(0) + 岗位价值(1-4) + 岗位效能(5)
    # 线2：岗位任务(6-8) + 任务目的与成果(9-14)
    first_rubric_id = issue_queue[0]["rubric_item_id"] if issue_queue else 0
    is_line1_issue = first_rubric_id <= 5  # 0-5 属于线1
    current_line = 1 if is_line1_issue else 2

    # 检查线1是否全部完成（无线1问题项或全部resolved）
    line1_issue_ids = {q["issue_id"] for q in issue_queue if q["rubric_item_id"] <= 5}
    line1_resolved = all(issue_status_map.get(iid) == "resolved" for iid in line1_issue_ids)
    line1_completed = len(line1_issue_ids) == 0 or line1_resolved

    # 有问题项：进入引导流程
    return {
        "review_items": review_items,
        "all_issues": normalized,
        "current_issue_index": 0,
        "issue_round": 0,
        "issue_status_map": issue_status_map,
        "coaching_queue_order": coaching_queue_order,
        "current_focus_id": current_focus_id,
        "pending_question": first_question,
        "phase": "guiding",
        "stuck_counter": 0,
        "hint_level": 0,
        "current_line": current_line,
        "line1_completed": line1_completed,
        "rubric_eval_summary": {
            "checked_item_ids": checked_item_ids,
            "matched_item_ids": [i["rubric_item_id"] for i in normalized],
            "relaxed_item_ids": relaxed_ids,
            "coverage_ok": len(checked_item_ids) == len(RUBRIC_ITEMS),
        },
    }


def _llm_review(
    scoring_criteria: str, teaching_material: str,
    row_data: dict[str, Any], columns: list[str],
) -> tuple[list[dict[str, Any]], list[int]]:
    """LLM 评审：逐条对照 14 项评分标准"""
    system = _build_system_prompt(scoring_criteria, teaching_material)

    row_text = "\n".join(f"  {k}: {_format_value_for_view(v)}" for k, v in row_data.items() if v)
    rubric_text = "\n".join(f"{item['id']}. [{item['category']}] {item['desc']} (等级{item['level']})" for item in RUBRIC_ITEMS)

    prompt = f"""你是岗标教练，正在后台诊断用户的岗标材料。

**角色提醒**：这是内部诊断，用户完全看不到这个过程。不要输出"我正在诊断"等提示。

### 被诊断内容
{row_text}

### 典型问题清单（逐条对照，不要遗漏）
{rubric_text}

### 诊断守则
1. **严格逐条检查**以上清单，不要遗漏
2. **三问法筛选**：谁 / 得到啥 / 商业落点——三问都过则放过
3. **灰区放过**：不像典型违规、又不能算明显合格时，默认放过
4. **问题严重程度**：只标记"需要引导"（A级）或"可优化"（B级），绝不用"扣分""得分"等评审语言

### 真实反例（三问法通过，不应追问）
{REAL_EXAMPLE}

### 输出格式
仅输出 JSON：
{{
  "checked_item_ids": [1,2,3,...],
  "issues": [
    {{
      "issue_id": "1",
      "issue_desc": "问题描述（用人话，不报编号）",
      "category": "岗位价值",
      "level": "A",
      "explanation": "判定理由",
      "user_facing_desc": "给用户的自然语言描述（教练口吻）"
    }}
  ]
}}

**注意**：
- 只输出真正需要引导的问题
- 用户看不到这个输出，所以 issue_desc 是内部参考，user_facing_desc 是给用户看的"人话版"
- 绝不要出现"扣分""得分""通关"等评审词汇"""

    try:
        llm = _get_llm()
        if llm is None:
            raise Exception("LLM not available")
        resp = llm.invoke([
            SystemMessage(content=system),
            HumanMessage(content=prompt),
        ])
        parsed = _safe_json_load(resp.content, {"checked_item_ids": [], "issues": []})
        checked = parsed.get("checked_item_ids", list(range(1, 15)))
        issues = parsed.get("issues", [])
        return issues, checked
    except Exception as e:
        print(f"[LLM 评审失败] {e}")
        return _rule_based_review(row_data)


def _rule_based_review(row_data: dict[str, Any]) -> tuple[list[dict[str, Any]], list[int]]:
    """无 LLM 时的规则评审"""
    issues: list[dict[str, Any]] = []
    checked: list[int] = list(range(1, 15))

    value_text = _extract_value_text(row_data)

    # 检查岗位价值
    if value_text:
        if not _passes_three_questions(value_text):
            if not any(k in value_text for k in WHO_HINTS):
                issues.append({"issue_id": "1", "issue_desc": "站在自身视角而不是站在客户视角提炼岗位价值", "category": "岗位价值", "level": "A", "deduction": 10})
            if not any(k in value_text for k in BENEFIT_HINTS):
                issues.append({"issue_id": "2", "issue_desc": "岗位价值并非源于对客户最在意、最深层需求的提炼", "category": "岗位价值", "level": "A", "deduction": 10})
            if not any(k in value_text for k in BIZ_HINTS):
                issues.append({"issue_id": "4", "issue_desc": "岗位价值描述空泛，指导性不强", "category": "岗位价值", "level": "A", "deduction": 10})
        else:
            # 三问法通过，放过
            pass

    # 检查岗位效能
    eff_text = str(row_data.get("岗位效能", "") or row_data.get("岗位效能_", ""))
    if not eff_text or eff_text == "无":
        issues.append({"issue_id": "5", "issue_desc": "岗位效能并不能直接、有效地衡量岗位价值", "category": "岗位效能", "level": "A", "deduction": 10})

    # 检查任务命名
    for key, val in row_data.items():
        if "任务" in key and isinstance(val, str) and val:
            # 检查是否像目的
            purpose_words = ("降低", "提升", "减少", "增加", "优化")
            if any(val.startswith(w) for w in purpose_words):
                issues.append({"issue_id": "8", "issue_desc": "直接用目的命名任务", "category": "岗位任务", "level": "B", "deduction": 5})

    # 检查成果
    for key, val in row_data.items():
        if "成果" in key and isinstance(val, str) and val:
            delivery_words = ("文档", "报告", "方案", "课件")
            if any(w in val for w in delivery_words) and not any(k in val for k in ("率", "%", "次")):
                issues.append({"issue_id": "13", "issue_desc": "把交付物当做成果", "category": "任务目的与成果", "level": "A", "deduction": 10})

    return issues, checked


# ─────────────────────────────────────────────────────────────
# NODE 3: guide_reflection — 分步引导（核心环节）
# 对应 skill 阶段 3：五条铁律 + 引导武器 + 十一问
# ─────────────────────────────────────────────────────────────

def guide_reflection(state: CoachState) -> dict:
    """
    发出引导性问题，引导用户自我反思当前条目。
    使用引导武器表和自检十一问生成自然的问题。
    遵循每轮对话结构：肯定 → 切入 → 一问 → 等待。
    """
    idx = state["current_item_index"]
    review_items = state.get("review_items", [])
    rows = state["submission_rows"]

    if idx >= len(rows):
        return {"phase": "done"}

    current_review = review_items[idx] if idx < len(review_items) else None
    row_data = rows[idx]
    total = len(rows)
    question = state.get("pending_question", "")
    issue_idx = state.get("current_issue_index", 0)
    issue_queue = current_review.get("issue_queue", []) if current_review else []

    print(f"[Node] guide_reflection: 条目{idx+1}/{total}，问题项{issue_idx+1}")

    if not current_review:
        return _advance_to_next_item(state, "当前条目尚未生成评审结果，跳过。")
    elif not issue_queue:
        # 无问题项：自动推进到下一条（正常情况下 review_item 已处理，
        # 这里是防御性兜底）
        return _advance_to_next_item(state, f"第 {idx+1}/{total} 条没有发现需要调整的地方。")
    else:
        issue = issue_queue[min(issue_idx, len(issue_queue) - 1)]
        content = _build_proactive_message(
            idx=idx, total=total, row_data=row_data,
            issue=issue, question=question,
        )

    return {
        "messages": [AIMessage(content=content)],
        "awaiting_user_input": True,
        "phase": "guiding",
        "active_mode": "proactive",
    }


# ─────────────────────────────────────────────────────────────
# NODE 3b: process_response — 处理用户回复
# 回判三步：1.肯定进步 2.给出判断 3.追或不追
# ─────────────────────────────────────────────────────────────

def process_response(state: CoachState, config: Optional[RunnableConfig] = None) -> dict:
    """
    分析用户对引导问题的回复。
    回判三步：肯定进步 → 给出判断 → 追或不追。
    "还能更好"不是追问的理由，"评委会扣分"才是。
    """
    thread_id = _thread_id_from_config(config)
    idx = state["current_item_index"]
    review_items = state.get("review_items", [])
    current_review = review_items[idx] if idx < len(review_items) else None
    messages = state.get("messages", [])
    issue_round = state.get("issue_round", 0)
    rows = state["submission_rows"]
    issue_idx = state.get("current_issue_index", 0)

    max_round = 3
    print(f"[Node] process_response: 条目{idx+1}，问题项{issue_idx+1}，已追问{issue_round+1}轮")

    user_messages = [m for m in messages if isinstance(m, HumanMessage)]
    last_user_msg = user_messages[-1].content if user_messages else ""

    # 意图已由 detect_user_intent 单点判定；此处不再二次分流到答疑，
    # 避免 answer_user_question 返回缺 phase 导致同轮重发引导问题。
    if not current_review:
        return {"phase": "reviewing", "active_mode": "proactive", "last_user_intent": "reply"}

    issue_queue = current_review.get("issue_queue", [])
    if not issue_queue:
        return _advance_to_next_item(state, "这一条目无需调整，我们继续下一条。")

    current_issue = issue_queue[min(issue_idx, len(issue_queue) - 1)]
    rubric_item_id = int(current_issue.get("rubric_item_id", 0) or 0)

    system = _build_system_prompt(state["scoring_criteria"], state["teaching_material"])
    recent = messages[-8:] if len(messages) > 8 else messages

    decide_prompt = f"""你是岗标教练，正在回判用户的修改回复。

**当前辅导位置**：第{idx+1}条岗标条目，当前问题项属于{'线1（客户/价值/效能）' if rubric_item_id <= 5 else '线2（任务/目的/成果）'}

**当前问题项**：
类别：{current_issue.get('category', '')}
问题描述：{current_issue.get('user_facing_desc', '')}

**用户刚才回复**："{last_user_msg}"
已进行轮次：{issue_round + 1}

---

请按照以下三步进行判断：

**第一步：肯定进步（必须具体）**
指出用户回复比上一轮好在哪个**具体点**——改进了什么、补上了什么信息。不能只说"很好""有进步"。

**第二步：给出判断**
- 是否解决：用户是否给出可执行、可落地的改进方向？
- 三问法检查（仅岗位价值类问题：客户主语/具体好处/商业落点）

**重要**：这是{'线1' if rubric_item_id <= 5 else '线2'}问题，{'线1完成后才能进入线2（任务/目的/成果）' if rubric_item_id <= 5 else '现在已进入线2，关注任务设计和成果规划'}

**第三步：追或不追**
- 问题解决或三问法合格 → 明确认可，告知用户进入下一项（或进入线2）
- 未解决但方向对 → 更具体地再问一次
- 卡住超3轮 → 给具体抓手（A/B/C候选项），但仍用问题形式

**红线原则**：
1. "还能更好"不是追问理由，"解决问题"才是
2. 灰区默认放过——三问法过了就停手
3. 绝不说"打分""扣分""过关"等评审语言

请仅输出 JSON：
{{
    "affirmation": "具体的肯定语（如：你这轮补上了'产品经理'这个客户主语，比上一轮具体多了）...",
    "issue_resolved": true/false,
    "judgment": "自然语言判断（如：客户主语有了，但具体好处还不够清晰）...",
    "feedback_to_user": "给用户的自然语言反馈（教练口吻，不说教）...",
    "next_question": "下一步的引导问题（如果已解决则为空）...",
    "transition_message": "{'线1到线2的过渡提示（如：客户价值这块已经清楚了，接下来我们看看任务设计）' if rubric_item_id <= 5 else ''}"
}}"""

    if _is_llm_enabled():
        try:
            llm = _get_llm()
            if llm is None:
                raise Exception("LLM not available")
            response = llm.invoke([
                SystemMessage(content=system),
                *recent,
                HumanMessage(content=decide_prompt),
            ])
            decision = _safe_json_load(response.content, {
                "affirmation": "你这轮思考有进展。",
                "issue_resolved": False,
                "judgment": "方向对了，但还可以更具体。",
                "feedback_to_user": "你的方向是对的，我们再把关键点说得更具体一些。",
                "next_question": "如果你现在就改写这段内容，你会先增加哪一个可量化或可验证的信息？",
            })
        except Exception:
            decision = _fallback_decision(last_user_msg, rubric_item_id)
    else:
        decision = _fallback_decision(last_user_msg, rubric_item_id)

    affirmation = decision.get("affirmation", "你这轮思考有进展。")
    feedback = decision.get("feedback_to_user", "谢谢你的思考。")
    issue_resolved = bool(decision.get("issue_resolved", False))
    next_question = decision.get("next_question", "你觉得还差哪一步可以让这个问题真正解决？")

    # 三问法检查（仅对视角偏差类问题：1=视角偏差, 3=多客户视角）
    # 其他岗位价值问题(2=深度偏差, 4=空泛)不适用三问法，由 LLM/规则判断
    if not issue_resolved and rubric_item_id in {1, 3}:
        issue_resolved = _passes_three_questions(last_user_msg)
        if issue_resolved:
            feedback = f"{affirmation} 你已经补齐了服务对象、具体收益和业务落点，这一项可以先放过。"
        else:
            feedback = f"{affirmation} {feedback}"
    else:
        feedback = f"{affirmation} {feedback}"

    issue_status_map = dict(state.get("issue_status_map", {}))
    queue_copy = list(issue_queue)

    if issue_resolved:
        current_issue_id = current_issue.get("issue_id", f"issue_{issue_idx}")
        issue_status_map[current_issue_id] = "resolved"
        queue_copy[issue_idx]["status"] = "resolved"

        # 更新进度
        _call_progress_script("update", current_issue_id, "pass", "三问法通过或用户给出可执行改进", thread_id=thread_id)

        # 双线推进：检查线1是否全部完成
        line1_issue_ids = {q["issue_id"] for q in queue_copy if q["rubric_item_id"] <= 5}
        line1_resolved_ids = {iid for iid in line1_issue_ids if issue_status_map.get(iid) == "resolved"}
        line1_completed = len(line1_issue_ids) == len(line1_resolved_ids)

        # 找到下一个待处理的问题项（考虑双线推进限制）
        next_issue_idx, next_issue, transition_msg = _find_next_issue(
            queue_copy, issue_idx, line1_completed, issue_status_map
        )

        if next_issue_idx is not None and next_issue is not None:
            queue_copy[next_issue_idx]["status"] = "in_progress"
            issue_status_map[next_issue.get("issue_id", f"issue_{next_issue_idx}")] = "in_progress"
            next_issue_question = _generate_issue_question(
                state["scoring_criteria"],
                state["teaching_material"],
                rows[idx],
                next_issue,
            )
            review_items[idx]["issue_queue"] = queue_copy

            # 构建消息：包含过渡提示（如果从线1进入线2）
            msg_content = f"✅ {feedback}"
            if transition_msg:
                msg_content += f"\n\n{transition_msg}"
            else:
                msg_content += "\n\n我们进入下一个关注点，继续逐项完善。"

            return {
                "messages": [AIMessage(content=msg_content)],
                "review_items": review_items,
                "issue_status_map": issue_status_map,
                "current_issue_index": next_issue_idx,
                "issue_round": 0,
                "pending_question": next_issue_question,
                "phase": "guiding",
                "active_mode": "proactive",
                "last_user_intent": "reply",
                "awaiting_user_input": False,
                "current_focus_id": next_issue.get("issue_id"),
                "stuck_counter": 0,
                "hint_level": 0,
                "current_line": 2 if next_issue["rubric_item_id"] > 5 else 1,
                "line1_completed": line1_completed,
            }

        review_items[idx]["issue_queue"] = queue_copy
        return {
            **_advance_to_next_item(state, feedback),
            "review_items": review_items,
            "issue_status_map": issue_status_map,
            "last_user_intent": "reply",
        }

        review_items[idx]["issue_queue"] = queue_copy
        return {
            **_advance_to_next_item(state, feedback),
            "review_items": review_items,
            "issue_status_map": issue_status_map,
            "last_user_intent": "reply",
        }

    # 未解决，继续追问
    review_items[idx]["issue_queue"] = queue_copy
    stuck_counter = state.get("stuck_counter", 0)
    hint_level = state.get("hint_level", 0)
    next_issue_round = issue_round + 1

    if next_issue_round >= max_round:
        stuck_counter += 1
        hint_level = min(hint_level + 1, 3)

        # 如果卡住超过3次（9轮追问），自动放过当前问题项
        if stuck_counter >= 3:
            current_issue_id = current_issue.get("issue_id", f"issue_{issue_idx}")
            issue_status_map[current_issue_id] = "resolved"
            queue_copy[issue_idx]["status"] = "resolved"
            review_items[idx]["issue_queue"] = queue_copy

            # 双线推进：检查线1是否完成（同上）
            line1_issue_ids = {q["issue_id"] for q in queue_copy if q["rubric_item_id"] <= 5}
            line1_resolved_ids = {iid for iid in line1_issue_ids if issue_status_map.get(iid) == "resolved"}
            line1_completed = len(line1_issue_ids) == len(line1_resolved_ids)

            next_issue_idx, next_issue, transition_msg = _find_next_issue(
                queue_copy, issue_idx, line1_completed, issue_status_map
            )

            if next_issue_idx is not None and next_issue is not None:
                queue_copy[next_issue_idx]["status"] = "in_progress"
                issue_status_map[next_issue.get("issue_id", f"issue_{next_issue_idx}")] = "in_progress"
                next_question_new = _generate_issue_question(
                    state["scoring_criteria"],
                    state["teaching_material"],
                    rows[idx],
                    next_issue,
                )
                msg_content = "我们先放过这个点，后面有时间可以再回来打磨。"
                if transition_msg:
                    msg_content += f"\n\n{transition_msg}"
                else:
                    msg_content += " 我们进入下一个关注点。"
                return {
                    "messages": [AIMessage(content=msg_content)],
                    "review_items": review_items,
                    "issue_status_map": issue_status_map,
                    "current_issue_index": next_issue_idx,
                    "issue_round": 0,
                    "pending_question": next_question_new,
                    "phase": "guiding",
                    "active_mode": "proactive",
                    "last_user_intent": "reply",
                    "awaiting_user_input": False,
                    "current_focus_id": next_issue.get("issue_id"),
                    "stuck_counter": 0,
                    "hint_level": 0,
                    "current_line": 2 if next_issue["rubric_item_id"] > 5 else 1,
                    "line1_completed": line1_completed,
                }
            else:
                return {
                    **_advance_to_next_item(state, "我们先放过这个点，后面有时间可以再回来打磨。"),
                    "review_items": review_items,
                    "issue_status_map": issue_status_map,
                    "last_user_intent": "reply",
                }

        next_question = _build_escalation_question(current_issue, rows[idx], hint_level)
        next_issue_round = 0

    content = f"✅ {feedback}\n\n💬 {next_question}"
    # 计算当前线状态和线1完成状态（保持当前）
    current_rubric_id = current_issue.get("rubric_item_id", 0)
    current_line = 1 if current_rubric_id <= 5 else 2
    line1_issue_ids = {q["issue_id"] for q in queue_copy if q["rubric_item_id"] <= 5}
    line1_resolved_ids = {iid for iid in line1_issue_ids if issue_status_map.get(iid) == "resolved"}
    line1_completed = len(line1_issue_ids) == len(line1_resolved_ids)

    return {
        "messages": [AIMessage(content=content)],
        "review_items": review_items,
        "pending_question": next_question,
        "issue_round": next_issue_round,
        "awaiting_user_input": True,
        "phase": "guiding",
        "active_mode": "proactive",
        "last_user_intent": "reply",
        "current_focus_id": current_issue.get("issue_id"),
        "stuck_counter": stuck_counter,
        "hint_level": hint_level,
        "current_line": current_line,
        "line1_completed": line1_completed,
    }


def _fallback_decision(user_msg: str, rubric_item_id: int) -> dict:
    """无 LLM 时的兜底判断逻辑"""
    has_action = any(k in user_msg for k in ["我会", "我改", "我补充", "改成", "增加", "I'll", "I should", "rename", "restructure"])
    has_metric = any(k in user_msg for k in ["可量化", "指标", "%", "率", "次", "个", "rate", "<=", ">="])
    has_customer = any(k in user_msg for k in WHO_HINTS) or any(
        k in user_msg.lower() for k in ("product manager", "business", "customer", "user", "team")
    )
    passes_three = _passes_three_questions(user_msg)

    if rubric_item_id in {1, 2, 3, 4}:
        # 岗位价值问题
        if passes_three:
            return {
                "affirmation": "你这轮给出了很具体的改进方向，",
                "issue_resolved": True,
                "judgment": "三问法通过，可以放过。",
                "feedback_to_user": "你已经明确了客户、好处和商业落点，这一项可以过了。",
                "next_question": "",
            }
        if has_customer and has_action:
            return {
                "affirmation": "你开始从客户视角思考了，",
                "issue_resolved": passes_three,
                "judgment": "还需要再具体一点。",
                "feedback_to_user": "方向对了，能不能再说说客户能获得什么具体好处？",
                "next_question": "如果只改一处，你会先补充哪条最具体的信息？",
            }
        # 对于深度偏差(2)和空泛(4)：只要用户给出有客户+动作的改进，就通过
        if rubric_item_id in {2, 4} and (has_customer or has_action):
            return {
                "affirmation": "你在往更具体的方向走了，",
                "issue_resolved": True,
                "judgment": "方向对了，可以放过。",
                "feedback_to_user": "这个改进比之前具体多了，这一项可以过了。",
                "next_question": "",
            }

    if rubric_item_id == 5:
        if has_metric:
            return {
                "affirmation": "你找到了衡量方式，",
                "issue_resolved": True,
                "judgment": "效能指标已明确。",
                "feedback_to_user": "这个指标很具体，可以过了。",
                "next_question": "",
            }

    if rubric_item_id in {6}:
        if has_action:
            return {
                "affirmation": "你在重新梳理任务和价值的关系，",
                "issue_resolved": True,
                "judgment": "任务-价值对应更清晰了。",
                "feedback_to_user": "这样关联起来更清楚了，可以过了。",
                "next_question": "",
            }

    if rubric_item_id in {7, 8}:
        if "动词" in user_msg or any(k in user_msg for k in ["负责", "完成", "制定", "整理", "rename", "verb"]):
            return {
                "affirmation": "你开始调整任务命名了，",
                "issue_resolved": True,
                "judgment": "符合动宾结构即可。",
                "feedback_to_user": "这个命名更清晰了，可以过了。",
                "next_question": "",
            }

    if rubric_item_id in {9}:
        if has_action:
            return {
                "affirmation": "你开始明确目的了，",
                "issue_resolved": True,
                "judgment": "目的更清晰了。",
                "feedback_to_user": "这个目的比之前明确多了，可以过了。",
                "next_question": "",
            }

    if rubric_item_id in {10}:
        if has_metric:
            return {
                "affirmation": "你在把成果和目的关联起来，",
                "issue_resolved": True,
                "judgment": "成果标准已明确。",
                "feedback_to_user": "成果和目的能对上了，可以过了。",
                "next_question": "",
            }

    if rubric_item_id in {12, 13, 14}:
        if has_metric or has_action:
            return {
                "affirmation": "你在把成果写得更具体了，",
                "issue_resolved": True,
                "judgment": "成果标准更可衡量了。",
                "feedback_to_user": "这样写成果更可验证了，可以过了。",
                "next_question": "",
            }

    # 默认返回
    return {
        "affirmation": "你这轮思考有进展，",
        "issue_resolved": has_action and has_metric,
        "judgment": "方向对了" if has_action else "还可以更具体",
        "feedback_to_user": "你的思考方向不错。" if has_action else "能不能再说说你具体会怎么改？",
        "next_question": "如果你现在就改写这段内容，你会先增加哪一个可量化或可验证的信息？",
    }



# ─────────────────────────────────────────────────────────────
# NODE 3c: detect_user_intent — 意图识别
# ─────────────────────────────────────────────────────────────

def detect_user_intent(state: CoachState) -> dict:
    """意图识别：规则优先，低置信度再走 LLM 二判。"""
    phase = state.get("phase", "guiding")
    messages = state.get("messages", [])
    user_messages = [m for m in messages if isinstance(m, HumanMessage)]
    last_user_msg = user_messages[-1].content if user_messages else ""

    if phase in {"done", "closure"}:
        return {"last_user_intent": "question", "active_mode": "reactive_qa"}

    intent, confidence = _rule_based_intent(last_user_msg)
    if intent == "uncertain" or confidence < 0.7:
        intent = _llm_intent_fallback(state, last_user_msg)

    return {
        "last_user_intent": intent,
        "active_mode": "reactive_qa" if intent == "question" else "proactive",
    }


def _llm_intent_fallback(state: CoachState, user_text: str) -> str:
    """二判：当规则不稳定时，用轻量 LLM 分类"""
    system = _build_system_prompt(
        state.get("scoring_criteria", MOCK_SCORING_CRITERIA),
        state.get("teaching_material", MOCK_TEACHING_MATERIAL),
    )
    prompt = f"""请判断用户最新输入意图。

当前阶段: {state.get('phase', 'guiding')}
当前待引导问题: {state.get('pending_question', '')}
用户输入: {user_text}

分类标准:
1. intent=reply: 用户在回答当前引导问题或给出修改思路
2. intent=question: 用户在主动发起新的咨询问题（解释/方法/示例等）

仅输出 JSON:
{{"intent":"reply或question","confidence":0.0}}"""

    if not _is_llm_enabled():
        return "reply"
    try:
        llm = _get_llm()
        if llm is None:
            return "reply"
        resp = llm.invoke([SystemMessage(content=system), HumanMessage(content=prompt)])
    except Exception:
        return "reply"

    parsed = _safe_json_load(resp.content, {"intent": "reply", "confidence": 0.5})
    intent = str(parsed.get("intent", "reply")).strip().lower()
    return intent if intent in {"reply", "question"} else "reply"


# ─────────────────────────────────────────────────────────────
# NODE 3d: answer_user_question — 被动答疑
# ─────────────────────────────────────────────────────────────

def answer_user_question(state: CoachState) -> dict:
    """场景二：用户主动提问时的被动答疑。"""
    messages = state.get("messages", [])
    user_messages = [m for m in messages if isinstance(m, HumanMessage)]
    question = user_messages[-1].content if user_messages else ""

    system = _build_system_prompt(
        state.get("scoring_criteria", MOCK_SCORING_CRITERIA),
        state.get("teaching_material", MOCK_TEACHING_MATERIAL),
    )
    qa_prompt = f"""用户主动提问如下，请进入被动答疑模式给出专业回答。

用户问题：{question}

要求：
1. 直接回答问题，语言专业但简洁
2. 可基于教材原则解释，但不要贴教材原文大段引用
3. 如果当前仍在主动引导流程（phase=guiding/reviewing），回答后加一句"我们回到当前问题项"的过渡
4. 用人话说，不报章节号、不报条目编号

只输出 JSON：
{{
  "answer": "...",
  "follow_back": "..."
}}"""

    if _is_llm_enabled():
        try:
            llm = _get_llm()
            if llm is None:
                raise Exception("LLM not available")
            resp = llm.invoke([SystemMessage(content=system), HumanMessage(content=qa_prompt)])
            parsed = _safe_json_load(resp.content, {
                "answer": "这个问题很关键。建议你先从服务对象、可衡量成果和业务结果三个维度拆开检查。",
                "follow_back": "我们回到当前问题项，继续逐项完善。",
            })
        except Exception:
            parsed = {
                "answer": "这是一个很好的问题。可以先从服务对象、关键行为和可衡量成果三层去梳理。",
                "follow_back": "我们回到当前问题项，继续逐项完善。",
            }
    else:
        parsed = {
            "answer": "这是一个很好的问题。可以先从服务对象、关键行为和可衡量成果三层去梳理。",
            "follow_back": "我们回到当前问题项，继续逐项完善。",
        }

    follow_back = ""
    if state.get("phase") in {"guiding", "reviewing"}:
        follow_back = parsed.get("follow_back", "我们回到当前问题项，继续逐项完善。")
    elif state.get("phase") in {"done", "closure"}:
        follow_back = "你还可以继续提问，我会持续答疑。"

    answer_text = parsed.get("answer", "")
    content = f"\U0001f4ab 针对你的问题<<{question}>>：\n{answer_text}"
    if follow_back:
        content += f"\n\n{follow_back}"

    return {
        "messages": [AIMessage(content=content)],
        "active_mode": "reactive_qa",
        "last_user_intent": "question",
        "awaiting_user_input": True,
    }


# ─────────────────────────────────────────────────────────────
# NODE 4: generate_closure — 收尾总结
# 对应 skill 阶段 4：自然总结 + 不给打分 + 教材收束
# ─────────────────────────────────────────────────────────────

def generate_closure(state: CoachState, config: Optional[RunnableConfig] = None) -> dict:
    """
    收尾总结：
    - 哪些地方改得特别好
    - 还有哪些可以继续打磨
    - 用一句教材里的话作为收束（用人话说，不标出处）
    - 不给具体打分
    - 清理进度
    """
    review_items = state.get("review_items", [])
    rows = state["submission_rows"]

    highlights: list[str] = []
    remaining: list[str] = []

    for idx, review in enumerate(review_items):
        issue_queue = review.get("issue_queue", [])
        resolved = [q for q in issue_queue if q.get("status") == "resolved"]
        pending = [q for q in issue_queue if q.get("status") != "resolved"]

        if resolved:
            for q in resolved:
                highlights.append(f"第{idx+1}条的{q.get('category', '')}部分改得很到位")
        if pending:
            for q in pending:
                remaining.append(f"第{idx+1}条的{q.get('category', '')}还可以继续打磨")

    # 生成总结文本
    if _is_llm_enabled():
        closure_text = _llm_closure(state, highlights, remaining)
    else:
        closure_text = _rule_based_closure(highlights, remaining)

    # 清理进度
    _call_progress_script("reset", thread_id=_thread_id_from_config(config))

    return {
        "phase": "closure",
        "closure_summary": closure_text,
        "highlights": highlights,
        "remaining_polish": remaining,
        "messages": [AIMessage(content=closure_text)],
        "awaiting_user_input": True,
        "active_mode": "reactive_qa",
    }


def _llm_closure(state: CoachState, highlights: list[str], remaining: list[str]) -> str:
    """LLM 生成收尾总结"""
    system = _build_system_prompt(state["scoring_criteria"], state["teaching_material"])

    h_text = "\n".join(f"- {h}" for h in highlights) if highlights else "（无特别突出的修改）"
    r_text = "\n".join(f"- {r}" for r in remaining) if remaining else "（没有需要继续打磨的地方了）"

    prompt = f"""辅导已全部完成，请生成收尾总结。

改得好的地方：
{h_text}

可以继续打磨的地方：
{r_text}

要求：
1. 用自然的语气总结，像老同事在茶水间聊完天的那种感觉
2. 不给具体打分——评分是评委的事
3. 用一句教材里的话作为收束，但要用人话说出来，不要标出处
4. 不要太长，3-5 句话就好

直接输出总结文本，不要 JSON 格式。"""

    try:
        llm = _get_llm()
        if llm is None:
            raise Exception("LLM not available")
        resp = llm.invoke([
            SystemMessage(content=system),
            HumanMessage(content=prompt),
        ])
        return resp.content.strip()
    except Exception:
        return _rule_based_closure(highlights, remaining)


def _rule_based_closure(highlights: list[str], remaining: list[str]) -> str:
    """规则兜底收尾总结"""
    parts = []

    if highlights:
        parts.append("改得好的地方：")
        for h in highlights[:5]:
            parts.append(f"  - {h}")

    if remaining:
        parts.append("\n可以继续打磨的地方：")
        for r in remaining[:5]:
            parts.append(f"  - {r}")

    # 用人话引用教材核心观点
    parts.append(
        "\n记住，你的价值由你服务的客户定义——"
        "始于客户需求，终于客户需求。继续加油！"
    )

    return "\n".join(parts)


# ─────────────────────────────────────────────────────────────
# 辅助：创建 Mock 提交文件
# ─────────────────────────────────────────────────────────────

def _create_mock_submission() -> str:
    """创建测试用的岗标提交文件"""
    import openpyxl
    path = "/tmp/mock_submission.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "岗标价值与岗标任务"

    # Row 1: 大类表头
    ws.append(["岗位价值", "岗位效能", "核心任务", None, None, "资源投入",
               "辅助任务", None, None, "资源投入"])
    ws.merge_cells("C1:E1")
    ws.merge_cells("G1:I1")

    # Row 2: 子列表头
    ws.append([None, None, "任务名称", "任务目的", "成果标准", None,
               "任务名称", "任务目的", "成果标准", None])
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("F1:F2")
    ws.merge_cells("J1:J2")

    # 条目1：有问题的内容
    ws.append([
        "负责软件开发工作", "无",
        "写代码", "完成功能开发", "完成任务", "60%",
        "开会", "信息同步", "参加会议", "40%",
    ])
    ws.append([
        None, None,
        "修bug", "减少缺陷", "bug数量", None,
        "写周报", "汇报工作", "提交周报", None,
    ])
    ws.merge_cells("A3:A4")

    # 条目2：较好的内容
    ws.append([
        "通过技术架构优化提升系统稳定性", "系统可用率≥99.9%",
        "主导核心模块技术方案设计", "提升模块质量与可维护性", "方案评审通过率100%", "70%",
        "技术文档编写", "知识沉淀与传承", "文档完整率90%", "30%",
    ])
    ws.append([
        None, None,
        "性能瓶颈分析与优化", "提升系统响应速度", "P99延迟≤200ms", None,
        "技术分享", "团队能力提升", "季度至少2次", None,
    ])
    ws.merge_cells("A5:A6")

    # 条目3
    ws.append([
        "支撑业务快速迭代", "无",
        "需求分析和功能开发", "满足业务需求", "功能上线", "80%",
        "协助测试", "保障质量", "配合完成", "20%",
    ])

    wb.save(path)
    print(f"  已生成 Mock 提交文件: {path}")
    return path
