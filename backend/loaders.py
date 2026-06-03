"""
文件加载工具：读取 xlsx / pdf → 文本摘要
"""
from __future__ import annotations
import json
import openpyxl
import pymupdf


# ─────────────────────────────────────────────────────────────
# XLSX 读取
# ─────────────────────────────────────────────────────────────

def _unmerge_worksheet(ws) -> list[list]:
    """
    读取 worksheet 并将合并单元格的值填充到所有被合并的位置。
    返回二维列表 data[row_idx][col_idx]。
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # 先读取所有单元格的原始值
    data: list[list] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        data.append(list(row))

    # 将合并区域的值填充到所有被合并的单元格
    for merge_range in ws.merged_cells.ranges:
        top_val = data[merge_range.min_row - 1][merge_range.min_col - 1]
        for r in range(merge_range.min_row - 1, merge_range.max_row):
            for c in range(merge_range.min_col - 1, merge_range.max_col):
                data[r][c] = top_val

    return data


def load_xlsx_as_text(path: str, max_rows: int = 200) -> tuple[list[str], list[dict], str]:
    """
    读取 xlsx，返回 (columns, rows, text_summary)
    text_summary 用于传给 LLM 做上下文。
    支持合并单元格：将合并区域的值填充到所有被合并的位置。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    data = _unmerge_worksheet(ws)
    wb.close()

    if not data:
        return [], [], "（表格为空）"

    header_row = data[0]
    raw_cols = [str(c).strip() if c is not None else f"列{i}" for i, c in enumerate(header_row)]

    # 处理重复列名：追加序号后缀
    seen: dict[str, int] = {}
    columns: list[str] = []
    for name in raw_cols:
        if name in seen:
            seen[name] += 1
            columns.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            columns.append(name)

    rows: list[dict] = []
    for i, raw_row in enumerate(data[1:]):
        if i >= max_rows:
            break
        row_dict = {columns[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(raw_row)}
        rows.append(row_dict)

    # 过滤掉所有行都为空的列
    non_empty_cols = [
        col for col in columns
        if any(r.get(col, "") for r in rows)
    ]
    if non_empty_cols:
        columns = non_empty_cols
        rows = [{col: r.get(col, "") for col in columns} for r in rows]

    # 过滤掉所有列都为空的行
    rows = [r for r in rows if any(v for v in r.values())]

    # 生成可读文本摘要
    lines = ["=== 表格内容 ===", f"列名：{' | '.join(columns)}", ""]
    for idx, r in enumerate(rows):
        cell_strs = [f"{k}: {v}" for k, v in r.items() if v]
        lines.append(f"第{idx+1}行：{'  '.join(cell_strs)}")

    return columns, rows, "\n".join(lines)


def load_scoring_criteria(path: str) -> str:
    """加载评分标准 xlsx，返回结构化文本"""
    _, _, text = load_xlsx_as_text(path, max_rows=500)
    return f"【评分标准】\n{text}"


def _deduplicate_names(names: list[str]) -> list[str]:
    """对列名列表去重，重复的追加序号。"""
    seen: dict[str, int] = {}
    result: list[str] = []
    for name in names:
        if name in seen:
            seen[name] += 1
            result.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            result.append(name)
    return result


def _detect_header_rows(data: list[list]) -> int:
    """
    检测表头行数。支持多级表头（如第 1 行大类、第 2 行子列名）。
    如果第二行包含典型子表头关键词或短文本，判定为双行表头。
    """
    if len(data) < 3:
        return 1

    row2_vals = [str(v).strip() if v is not None else "" for v in data[1]]

    # 检查第二行是否含有典型的子表头关键词
    sub_header_keywords = {"任务名称", "任务目的", "成果标准", "名称", "目的",
                           "成果", "投入比例", "关键行为", "输出成果"}
    if any(v in sub_header_keywords for v in row2_vals):
        return 2

    # 检查第一行是否有重复值（合并单元格展开后的特征）且第二行有多个短文本
    row1_nonempty = [str(v).strip() for v in data[0] if v is not None and str(v).strip()]
    row2_nonempty = [v for v in row2_vals if v]
    if len(row1_nonempty) != len(set(row1_nonempty)) and len(row2_nonempty) >= 3:
        avg_len = sum(len(v) for v in row2_nonempty) / max(len(row2_nonempty), 1)
        if avg_len <= 8:
            return 2

    return 1


def _build_composite_columns(header_rows: list[list]) -> list[str]:
    """
    从多行表头构建复合列名。
    如第一行 "核心任务" 第二行 "任务名称" → "核心任务_任务名称"。
    如果上下行值相同（如 "岗位价值"/"岗位价值"），只保留一个。
    """
    if len(header_rows) == 1:
        row = header_rows[0]
        raw = [str(v).strip() if v is not None else f"列{i}" for i, v in enumerate(row)]
        return _deduplicate_names(raw)

    num_cols = max(len(row) for row in header_rows)
    columns: list[str] = []

    for col_idx in range(num_cols):
        parts: list[str] = []
        for row in header_rows:
            val = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
            if val and val not in parts:
                parts.append(val)
        name = "_".join(parts) if parts else f"列{col_idx}"
        columns.append(name)

    return _deduplicate_names(columns)


def _group_rows_by_value(
    flat_columns: list[str],
    flat_rows: list[dict],
    resource_indices: list[int],
) -> tuple[list[str], list[dict]]:
    """
    将扁平表格按 "资源投入" 列拆分为左右两部分，并按岗位价值分组。

    返回 (top_columns, grouped_rows)，每个 grouped_row 格式：
    {
        "岗位价值": "...",
        "岗位效能": "...",
        "核心任务": [{"任务名称": "...", "任务目的": "...", "成果标准": "..."}],
        "核心任务_资源投入": "60%",
        "辅助任务": [{"任务名称": "...", "任务目的": "...", "成果标准": "..."}],
        "辅助任务_资源投入": "40%",
    }
    """
    first_res = resource_indices[0]
    second_res = resource_indices[1]

    # 识别共享列（岗位价值、岗位效能等，在核心任务列之前）
    shared_cols: list[str] = []
    core_start = 0
    for i, col in enumerate(flat_columns):
        if i >= first_res:
            break
        if "任务" in col:
            core_start = i
            break
        shared_cols.append(col)
    else:
        core_start = len(shared_cols)

    core_task_cols = flat_columns[core_start:first_res]
    core_resource_col = flat_columns[first_res]
    aux_task_cols = flat_columns[first_res + 1:second_res]
    aux_resource_col = flat_columns[second_res]

    # 提取子列名（去掉 "核心任务_"/"辅助任务_" 等前缀）
    def strip_prefix(col_name: str) -> str:
        if "_" in col_name:
            return col_name.split("_", 1)[1]
        return col_name

    core_sub_cols = [strip_prefix(c) for c in core_task_cols]
    aux_sub_cols = [strip_prefix(c) for c in aux_task_cols]

    # 按岗位价值分组
    value_col = shared_cols[0] if shared_cols else None
    groups: list[dict] = []
    current_group: dict | None = None

    for row in flat_rows:
        val = row.get(value_col, "") if value_col else ""

        # 新组：岗位价值非空且与当前组不同
        if val and (current_group is None or val != current_group.get(value_col, "")):
            if current_group is not None:
                groups.append(current_group)
            current_group = {col: row[col] for col in shared_cols}
            current_group["核心任务"] = []
            current_group["核心任务_资源投入"] = ""
            current_group["辅助任务"] = []
            current_group["辅助任务_资源投入"] = ""

        if current_group is None:
            current_group = {col: "" for col in shared_cols}
            current_group["核心任务"] = []
            current_group["核心任务_资源投入"] = ""
            current_group["辅助任务"] = []
            current_group["辅助任务_资源投入"] = ""

        # 添加核心任务
        core_task = {sub: row.get(col, "") for col, sub in zip(core_task_cols, core_sub_cols)}
        if any(v for v in core_task.values()):
            current_group["核心任务"].append(core_task)

        # 添加辅助任务
        aux_task = {sub: row.get(col, "") for col, sub in zip(aux_task_cols, aux_sub_cols)}
        if any(v for v in aux_task.values()):
            current_group["辅助任务"].append(aux_task)

        # 资源投入（取第一个非空值）
        if row.get(core_resource_col) and not current_group["核心任务_资源投入"]:
            current_group["核心任务_资源投入"] = row[core_resource_col]
        if row.get(aux_resource_col) and not current_group["辅助任务_资源投入"]:
            current_group["辅助任务_资源投入"] = row[aux_resource_col]

    if current_group is not None:
        groups.append(current_group)

    # 构建顶层列名描述
    top_columns = list(shared_cols)
    if core_task_cols:
        top_columns.append("核心任务")
        top_columns.append("核心任务_资源投入")
    if aux_task_cols:
        top_columns.append("辅助任务")
        top_columns.append("辅助任务_资源投入")

    return top_columns, groups


def _build_submission_summary(
    columns: list[str],
    rows: list[dict],
    shared_cols: list[str],
    core_sub_cols: list[str],
    aux_sub_cols: list[str],
    is_split: bool,
) -> str:
    """生成结构化文本摘要。"""
    if is_split:
        lines = [
            "=== 表格结构 ===",
            '该表以"资源投入"列为界，分为左右两部分：',
            "",
            f"【左表 - 核心任务侧】共享列：{' | '.join(shared_cols)}；任务子列：{' | '.join(core_sub_cols)}",
            f"【右表 - 辅助任务侧】任务子列：{' | '.join(aux_sub_cols)}",
            "",
            f"共识别到 {len(rows)} 个岗位价值条目：",
            "",
        ]
        for idx, group in enumerate(rows):
            lines.append(f"--- 条目 {idx + 1} ---")
            for col in shared_cols:
                val = group.get(col, "")
                if val:
                    lines.append(f"  {col}：{val}")

            core_tasks = group.get("核心任务", [])
            core_res = group.get("核心任务_资源投入", "")
            if core_tasks:
                lines.append(f"  核心任务（资源投入：{core_res}）：")
                for i, task in enumerate(core_tasks, 1):
                    task_str = " | ".join(f"{k}: {v}" for k, v in task.items() if v)
                    lines.append(f"    {i}. {task_str}")

            aux_tasks = group.get("辅助任务", [])
            aux_res = group.get("辅助任务_资源投入", "")
            if aux_tasks:
                lines.append(f"  辅助任务（资源投入：{aux_res}）：")
                for i, task in enumerate(aux_tasks, 1):
                    task_str = " | ".join(f"{k}: {v}" for k, v in task.items() if v)
                    lines.append(f"    {i}. {task_str}")
            lines.append("")
        return "\n".join(lines)
    else:
        lines = ["=== 表格内容 ===", f"列名：{' | '.join(columns)}", ""]
        for idx, r in enumerate(rows):
            cell_strs = [f"{k}: {v}" for k, v in r.items() if v]
            lines.append(f"第{idx + 1}行：{'  '.join(cell_strs)}")
        return "\n".join(lines)


def load_submission(path: str) -> tuple[list[str], list[dict], str]:
    """
    加载用户提交的岗标表格，智能处理多级表头和左右分表结构。

    以"资源投入"列为界，将表格拆分为：
    - 左表（核心任务侧）：岗位价值 + 岗位效能 + 核心任务详情 + 资源投入
    - 右表（辅助任务侧）：辅助任务详情 + 资源投入

    支持两行表头（如第一行"核心任务"横跨多列，第二行是子列名）。
    多行数据按岗位价值分组，同一岗位价值下的多条任务归入同一条目。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    data = _unmerge_worksheet(ws)
    wb.close()

    if not data or len(data) < 2:
        return [], [], "（表格为空）"

    # 1. 检测表头行数
    num_header_rows = _detect_header_rows(data)

    # 2. 构建复合列名
    flat_columns = _build_composite_columns(data[:num_header_rows])

    # 3. 提取数据行
    flat_rows: list[dict] = []
    for raw_row in data[num_header_rows:]:
        row_dict = {}
        for j, col in enumerate(flat_columns):
            val = str(raw_row[j]).strip() if j < len(raw_row) and raw_row[j] is not None else ""
            row_dict[col] = val
        flat_rows.append(row_dict)

    # 过滤空行
    flat_rows = [r for r in flat_rows if any(v for v in r.values())]

    # 过滤全空列
    non_empty_cols = [col for col in flat_columns if any(r.get(col, "") for r in flat_rows)]
    flat_columns = non_empty_cols
    flat_rows = [{col: r.get(col, "") for col in flat_columns} for r in flat_rows]

    # 4. 查找"资源投入"列，判断是否为左右分表结构
    resource_indices = [i for i, col in enumerate(flat_columns) if "资源投入" in col]

    if len(resource_indices) >= 2:
        # 左右分表结构：按岗位价值分组
        columns, rows = _group_rows_by_value(flat_columns, flat_rows, resource_indices)

        # 提取子列名用于摘要
        first_res = resource_indices[0]
        shared_cols = [col for col in flat_columns[:first_res] if "任务" not in col]
        core_start = next((i for i, col in enumerate(flat_columns) if "任务" in col and i < first_res), first_res)
        core_sub_cols = [col.split("_", 1)[1] if "_" in col else col for col in flat_columns[core_start:first_res]]
        aux_sub_cols = [col.split("_", 1)[1] if "_" in col else col for col in flat_columns[first_res + 1:resource_indices[1]]]

        text = _build_submission_summary(columns, rows, shared_cols, core_sub_cols, aux_sub_cols, is_split=True)
    else:
        # 无分表结构，退化为扁平格式
        columns = flat_columns
        rows = flat_rows
        text = _build_submission_summary(columns, rows, [], [], [], is_split=False)

    return columns, rows, f"【用户提交的岗标表格】\n{text}"


# ─────────────────────────────────────────────────────────────
# PDF 读取
# ─────────────────────────────────────────────────────────────

def load_pdf_as_text(path: str, max_chars: int = 20000) -> str:
    """读取 PDF，返回纯文本（截断至 max_chars）"""
    doc = pymupdf.open(path)
    pages_text: list[str] = []
    total = 0
    for page in doc:
        t = page.get_text()
        pages_text.append(t)
        total += len(t)
        if total >= max_chars:
            break
    doc.close()
    full = "\n".join(pages_text)
    if len(full) > max_chars:
        full = full[:max_chars] + "\n…（已截断）"
    return f"【岗标教材】\n{full}"


# ─────────────────────────────────────────────────────────────
# Mock 数据（当真实文件不存在时使用，方便测试）
# ─────────────────────────────────────────────────────────────

MOCK_SCORING_CRITERIA = """【评分标准（示例）】
维度1 - 岗标价值描述（权重20%）：
  5分：价值清晰、与公司战略强关联、可量化
  3分：价值有描述但缺乏量化指标
  1分：价值描述模糊或与业务脱节

维度2 - 岗标任务完整性（权重20%）：
  5分：任务覆盖全面、优先级清晰、有时间维度
  3分：任务有列举但缺少优先级
  1分：任务描述过于宽泛

维度3 - 关键行为可观测性（权重20%）：
  5分：行为描述具体可观测、有情境+行动+结果
  3分：行为有描述但缺少结果层面
  1分：行为描述抽象无法评估

维度4 - 输出成果可衡量性（权重20%）：
  5分：成果有明确度量指标（数量/质量/时效）
  3分：成果有描述但度量模糊
  1分：成果无法衡量

维度5 - 列间逻辑一致性（权重20%）：
  5分：价值→任务→行为→成果逻辑链完整
  3分：部分逻辑链断裂
  1分：各列内容互相矛盾或无关联
"""

MOCK_TEACHING_MATERIAL = """【岗标教材核心要点（示例）】

§1 什么是岗标价值
岗标价值是岗位存在的根本理由，回答"这个岗位为什么被需要"。
好的岗标价值应当：
- 与公司战略或部门OKR形成呼应
- 描述对业务/用户/团队产生的影响而非工作内容本身
- 尽量包含可量化的方向（如"通过技术手段降低系统故障率"）

§2 岗标任务与价值的关系
岗标任务是实现岗标价值的"路径"。
常见错误：任务只写日常工作，没有体现价值方向。
正确写法：每项任务应能回答"这个任务如何支撑岗标价值"。

§3 关键行为的写法
关键行为采用"情境-行动-结果"三段式：
  情境：在什么场景下
  行动：做了什么具体行动
  结果：产生了什么可观测的结果
示例（差）：积极参与代码评审
示例（好）：在需求评审阶段主动识别技术风险点，输出书面风险清单，并推动相关方形成一致解决方案

§4 输出成果的度量设计
输出成果需满足"SMART"原则：
- Specific（具体）
- Measurable（可衡量）
- Achievable（可达到）
- Relevant（与任务相关）
- Time-bound（有时间基准）

§5 表格各列的关联关系
岗标价值 → 岗标任务 → 关键行为 → 输出成果
这是一条完整的"价值-路径-行动-结果"逻辑链。
填写时应从左到右保持逻辑一致：
- 输出成果的度量指标应能证明关键行为已发生
- 关键行为应是完成岗标任务的必要动作
- 岗标任务应是实现岗标价值的充分路径
"""
REQUIRED_BLOCKS = ["岗位价值", "岗位任务"]


def validate_structure(wb):
    """校验 workbook 是否包含必需区块。"""
    sheet_names = wb.sheetnames
    sheet_hit = {b: any(b in s for s in sheet_names) for b in REQUIRED_BLOCKS}

    header_blocks = set()
    matched_sheet = None
    for s in sheet_names:
        ws = wb[s]
        headers = ""
        for row in ws.iter_rows(values_only=True, max_row=5):
            headers += " " + " ".join(str(c) for c in row if c)
        hits = [b for b in REQUIRED_BLOCKS if b in headers]
        if hits:
            matched_sheet = s
            header_blocks.update(hits)

    detected = sorted(header_blocks | {b for b, v in sheet_hit.items() if v})
    missing = [b for b in REQUIRED_BLOCKS if not sheet_hit[b] and b not in header_blocks]
    return missing, {
        "sheets_found": sheet_names,
        "matched_sheet": matched_sheet,
        "blocks_detected": detected,
        "missing": missing,
    }


def _is_combined_sheet(sheet_name):
    """判断 sheet 名是否同时包含两个区块关键字（即单 sheet 标准模板）。"""
    return all(b in sheet_name for b in REQUIRED_BLOCKS)


def extract_structured(wb):
    """结构化提取：按客户分组，把原始行解析为可评审的结构。"""
    result = {"clients": [], "auxiliary_tasks": []}
    sheet_names = wb.sheetnames

    # 查找单 sheet 标准模板（sheet 名同时含两个区块关键字）
    combined = None
    for s in sheet_names:
        if _is_combined_sheet(s):
            combined = s
            break

    # 查找单 sheet 但 sheet 名不含关键字的（用列头检测）
    if not combined:
        for s in sheet_names:
            ws = wb[s]
            headers = ""
            for row in ws.iter_rows(values_only=True, max_row=5):
                headers += " " + " ".join(str(c) for c in row if c)
            if all(b in headers for b in REQUIRED_BLOCKS):
                combined = s
                break

    if combined:
        rows = _read_rows(wb[combined])
        _parse_combined_sheet(rows, result)
        return result

    # 多 sheet 模式：按 sheet 名分别读取
    for block in REQUIRED_BLOCKS:
        for s in sheet_names:
            if block in s and not _is_combined_sheet(s):
                rows = _read_rows(wb[s])
                if block == "岗位价值":
                    _parse_value_sheet(rows, result)
                else:
                    _parse_task_sheet(rows, result)
                break

    return result


def _read_rows(ws):
    """读取 sheet 所有非空行。"""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            rows.append([str(c).strip() if c is not None else "" for c in row])
    return rows


# 标准模板列映射
COL = {"name": 0, "client": 1, "value": 2, "efficiency": 3,
       "resource": 4, "task_type": 5, "task_name": 6,
       "task_purpose_result": 7, "guide": 8}


def _parse_combined_sheet(rows, result):
    """解析单 sheet 标准模板，按客户分组提取价值和任务。"""
    current_client = None
    current_task_type = "核心任务"

    for row in rows:
        padded = row + [""] * (max(COL.values()) + 1 - len(row))

        # 跳过表头行
        if "岗位名称" in padded[0] or "岗位名称" in padded[COL["name"]]:
            continue

        # 检测客户切换（B 列非空）
        client_val = padded[COL["client"]].strip()
        if client_val:
            current_client = {
                "name": client_val,
                "position_value": padded[COL["value"]].strip(),
                "efficiency": padded[COL["efficiency"]].strip(),
                "tasks": []
            }
            result["clients"].append(current_client)

        # 检测任务类型切换
        task_type_val = padded[COL["task_type"]].strip()
        if task_type_val:
            current_task_type = task_type_val

        # 提取任务
        task_name = padded[COL["task_name"]].strip()
        task_pr = padded[COL["task_purpose_result"]].strip()
        guide = padded[COL["guide"]].strip()

        if task_name:
            task = {
                "type": current_task_type,
                "name": task_name,
                "purpose_and_result": task_pr,
                "guide": guide
            }

            if current_task_type == "辅助任务":
                result["auxiliary_tasks"].append(task)
            elif current_client:
                current_client["tasks"].append(task)


def _parse_value_sheet(rows, result):
    """多 sheet 模式：解析岗位价值 sheet。"""
    for row in rows:
        padded = row + [""] * 10
        if "岗位名称" in padded[0]:
            continue
        client_val = padded[1].strip()
        if client_val:
            result["clients"].append({
                "name": client_val,
                "position_value": padded[2].strip(),
                "efficiency": padded[3].strip(),
                "tasks": []
            })


def _parse_task_sheet(rows, result):
    """多 sheet 模式：解析岗位任务 sheet。"""
    current_client = None
    current_task_type = "核心任务"
    client_idx = 0

    for row in rows:
        padded = row + [""] * 10
        if "岗位名称" in padded[0]:
            continue

        if padded[1].strip():
            # 客户切换
            if client_idx < len(result["clients"]):
                current_client = result["clients"][client_idx]
                client_idx += 1

        task_type_val = padded[5].strip()
        if task_type_val:
            current_task_type = task_type_val

        task_name = padded[6].strip()
        if task_name:
            task = {
                "type": current_task_type,
                "name": task_name,
                "purpose_and_result": padded[7].strip(),
                "guide": padded[8].strip()
            }
            if current_task_type == "辅助任务":
                result["auxiliary_tasks"].append(task)
            elif current_client:
                current_client["tasks"].append(task)


def main(path: str) -> int:
    p = Path(path)
    if not p.exists():
        print(json.dumps({"ok": False, "error": f"文件不存在: {path}"}, ensure_ascii=False))
        return 1

    do_extract = "--extract" in sys.argv
    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"ok": False, "error": "需要安装 openpyxl: pip install openpyxl"}, ensure_ascii=False))
        return 1

    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        print(json.dumps({"ok": False, "error": f"无法打开 Excel: {e}"}, ensure_ascii=False))
        return 1

    missing, info = validate_structure(wb)

    result = {"ok": len(missing) == 0, "file": str(p)}
    result.update(info)

    if do_extract and not missing:
        result["data"] = extract_structured(wb)

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if not missing else 2


if __name__ == "__main__":
    if len(sys.argv) < 2 or sys.argv[1] in ("--help", "-h"):
        print("用法: python validate_sheets.py <xlsx_path> [--extract]", file=sys.stderr)
        sys.exit(0)
    sys.exit(main(sys.argv[1]))
