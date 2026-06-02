#!/usr/bin/env python3
"""验证并提取"岗位价值与岗位任务"Excel 内容。

用法:
    python validate_sheets.py <xlsx_path> [--extract]

选项:
    --extract  结构化提取内容，按客户分组输出

退出码:
    0 = 通过
    1 = 文件不存在 / 无法打开
    2 = 缺少必需区块
"""
import sys
import json
from pathlib import Path

REQUIRED_BLOCKS = ["岗位价值", "岗位任务"]


def validate_structure(wb):
    """校验 workbook 是否包含必需区块。"""
    sheet_names = wb.sheetnames
    sheet_hit = {b: any(b in s for s in sheet_names) for b in REQUIRED_BLOCKS}

    header_blocks = set()
    matched_sheet = None
    for s in sheet_names:
        ws = wb[s]
        headers = ""
        for row in ws.iter_rows(values_only=True, max_row=5):
            headers += " " + " ".join(str(c) for c in row if c)
        hits = [b for b in REQUIRED_BLOCKS if b in headers]
        if hits:
            matched_sheet = s
            header_blocks.update(hits)

    detected = sorted(header_blocks | {b for b, v in sheet_hit.items() if v})
    missing = [b for b in REQUIRED_BLOCKS if not sheet_hit[b] and b not in header_blocks]
    return missing, {
        "sheets_found": sheet_names,
        "matched_sheet": matched_sheet,
        "blocks_detected": detected,
        "missing": missing,
    }


def _is_combined_sheet(sheet_name):
    """判断 sheet 名是否同时包含两个区块关键字（即单 sheet 标准模板）。"""
    return all(b in sheet_name for b in REQUIRED_BLOCKS)


def extract_structured(wb):
    """结构化提取：按客户分组，把原始行解析为可评审的结构。"""
    result = {"clients": [], "auxiliary_tasks": []}
    sheet_names = wb.sheetnames

    # 查找单 sheet 标准模板（sheet 名同时含两个区块关键字）
    combined = None
    for s in sheet_names:
        if _is_combined_sheet(s):
            combined = s
            break

    # 查找单 sheet 但 sheet 名不含关键字的（用列头检测）
    if not combined:
        for s in sheet_names:
            ws = wb[s]
            headers = ""
            for row in ws.iter_rows(values_only=True, max_row=5):
                headers += " " + " ".join(str(c) for c in row if c)
            if all(b in headers for b in REQUIRED_BLOCKS):
                combined = s
                break

    if combined:
        rows = _read_rows(wb[combined])
        _parse_combined_sheet(rows, result)
        return result

    # 多 sheet 模式：按 sheet 名分别读取
    for block in REQUIRED_BLOCKS:
        for s in sheet_names:
            if block in s and not _is_combined_sheet(s):
                rows = _read_rows(wb[s])
                if block == "岗位价值":
                    _parse_value_sheet(rows, result)
                else:
                    _parse_task_sheet(rows, result)
                break

    return result


def _read_rows(ws):
    """读取 sheet 所有非空行。"""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            rows.append([str(c).strip() if c is not None else "" for c in row])
    return rows


# 标准模板列映射
COL = {"name": 0, "client": 1, "value": 2, "efficiency": 3,
       "resource": 4, "task_type": 5, "task_name": 6,
       "task_purpose_result": 7, "guide": 8}


def _parse_combined_sheet(rows, result):
    """解析单 sheet 标准模板，按客户分组提取价值和任务。"""
    current_client = None
    current_task_type = "核心任务"

    for row in rows:
        padded = row + [""] * (max(COL.values()) + 1 - len(row))

        # 跳过表头行
        if "岗位名称" in padded[0] or "岗位名称" in padded[COL["name"]]:
            continue

        # 检测客户切换（B 列非空）
        client_val = padded[COL["client"]].strip()
        if client_val:
            current_client = {
                "name": client_val,
                "position_value": padded[COL["value"]].strip(),
                "efficiency": padded[COL["efficiency"]].strip(),
                "tasks": []
            }
            result["clients"].append(current_client)

        # 检测任务类型切换
        task_type_val = padded[COL["task_type"]].strip()
        if task_type_val:
            current_task_type = task_type_val

        # 提取任务
        task_name = padded[COL["task_name"]].strip()
        task_pr = padded[COL["task_purpose_result"]].strip()
        guide = padded[COL["guide"]].strip()

        if task_name:
            task = {
                "type": current_task_type,
                "name": task_name,
                "purpose_and_result": task_pr,
                "guide": guide
            }

            if current_task_type == "辅助任务":
                result["auxiliary_tasks"].append(task)
            elif current_client:
                current_client["tasks"].append(task)


def _parse_value_sheet(rows, result):
    """多 sheet 模式：解析岗位价值 sheet。"""
    for row in rows:
        padded = row + [""] * 10
        if "岗位名称" in padded[0]:
            continue
        client_val = padded[1].strip()
        if client_val:
            result["clients"].append({
                "name": client_val,
                "position_value": padded[2].strip(),
                "efficiency": padded[3].strip(),
                "tasks": []
            })


def _parse_task_sheet(rows, result):
    """多 sheet 模式：解析岗位任务 sheet。"""
    current_client = None
    current_task_type = "核心任务"
    client_idx = 0

    for row in rows:
        padded = row + [""] * 10
        if "岗位名称" in padded[0]:
            continue

        if padded[1].strip():
            # 客户切换
            if client_idx < len(result["clients"]):
                current_client = result["clients"][client_idx]
                client_idx += 1

        task_type_val = padded[5].strip()
        if task_type_val:
            current_task_type = task_type_val

        task_name = padded[6].strip()
        if task_name:
            task = {
                "type": current_task_type,
                "name": task_name,
                "purpose_and_result": padded[7].strip(),
                "guide": padded[8].strip()
            }
            if current_task_type == "辅助任务":
                result["auxiliary_tasks"].append(task)
            elif current_client:
                current_client["tasks"].append(task)


def main(path: str) -> int:
    p = Path(path)
    if not p.exists():
        print(json.dumps({"ok": False, "error": f"文件不存在: {path}"}, ensure_ascii=False))
        return 1

    do_extract = "--extract" in sys.argv
    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"ok": False, "error": "需要安装 openpyxl: pip install openpyxl"}, ensure_ascii=False))
        return 1

    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        print(json.dumps({"ok": False, "error": f"无法打开 Excel: {e}"}, ensure_ascii=False))
        return 1

    missing, info = validate_structure(wb)

    result = {"ok": len(missing) == 0, "file": str(p)}
    result.update(info)

    if do_extract and not missing:
        result["data"] = extract_structured(wb)

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if not missing else 2


if __name__ == "__main__":
    if len(sys.argv) < 2 or sys.argv[1] in ("--help", "-h"):
        print("用法: python validate_sheets.py <xlsx_path> [--extract]", file=sys.stderr)
        sys.exit(0)
    sys.exit(main(sys.argv[1]))
